"""
Report Generators - PDF and Excel Report Generation
"""
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from django.utils import timezone
from inventory.models import InventoryItem, StockTransaction
from events.models import Event
import os


def get_report_directory():
    """Get or create reports directory"""
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    return reports_dir


def generate_inventory_status_pdf(user, category=None):
    """Generate Inventory Status PDF Report"""
    from reportlab.pdfgen import canvas
    
    reports_dir = get_report_directory()
    filename = f'inventory_status_{user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    file_path = os.path.join(reports_dir, filename)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("Inventory Status Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Date
    date_text = Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 0.2*inch))
    
    # Get inventory items
    items = InventoryItem.objects.filter(is_active=True)
    if category:
        items = items.filter(category=category)
    
    # Table data
    data = [['Name', 'SKU', 'Category', 'Current Stock', 'Min Stock', 'Status']]
    
    for item in items:
        status = 'Low Stock' if item.is_low_stock else 'OK'
        data.append([
            item.name,
            item.sku,
            item.get_category_display(),
            f"{item.current_stock} {item.unit}",
            f"{item.minimum_stock} {item.unit}",
            status
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return file_path


def generate_inventory_status_excel(user, category=None):
    """Generate Inventory Status Excel Report"""
    reports_dir = get_report_directory()
    filename = f'inventory_status_{user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    file_path = os.path.join(reports_dir, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Status"
    
    # Headers
    headers = ['Name', 'SKU', 'Category', 'Current Stock', 'Min Stock', 'Status', 'Stock Value']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Get inventory items
    items = InventoryItem.objects.filter(is_active=True)
    if category:
        items = items.filter(category=category)
    
    # Add data
    for item in items:
        status = 'Low Stock' if item.is_low_stock else 'OK'
        ws.append([
            item.name,
            item.sku,
            item.get_category_display(),
            f"{item.current_stock} {item.unit}",
            f"{item.minimum_stock} {item.unit}",
            status,
            f"£{item.stock_value:.2f}"
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(file_path)
    return file_path


def generate_stock_summary_pdf(user, start_date=None, end_date=None):
    """Generate Stock Summary PDF Report"""
    reports_dir = get_report_directory()
    filename = f'stock_summary_{user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    file_path = os.path.join(reports_dir, filename)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Stock Summary Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Get transactions
    transactions = StockTransaction.objects.all()
    if start_date:
        transactions = transactions.filter(created_at__gte=start_date)
    if end_date:
        transactions = transactions.filter(created_at__lte=end_date)
    
    data = [['Date', 'Item', 'Type', 'Quantity', 'Previous Stock', 'New Stock']]
    
    for transaction in transactions[:100]:  # Limit to 100 for PDF
        data.append([
            transaction.created_at.strftime('%Y-%m-%d'),
            transaction.inventory_item.name,
            transaction.get_transaction_type_display(),
            str(transaction.quantity),
            str(transaction.previous_stock),
            str(transaction.new_stock)
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return file_path


def generate_stock_summary_excel(user, start_date=None, end_date=None):
    """Generate Stock Summary Excel Report"""
    reports_dir = get_report_directory()
    filename = f'stock_summary_{user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    file_path = os.path.join(reports_dir, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Summary"
    
    headers = ['Date', 'Item', 'SKU', 'Type', 'Quantity', 'Previous Stock', 'New Stock', 'Performed By']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    transactions = StockTransaction.objects.all()
    if start_date:
        transactions = transactions.filter(created_at__gte=start_date)
    if end_date:
        transactions = transactions.filter(created_at__lte=end_date)
    
    for transaction in transactions:
        ws.append([
            transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            transaction.inventory_item.name,
            transaction.inventory_item.sku,
            transaction.get_transaction_type_display(),
            transaction.quantity,
            transaction.previous_stock,
            transaction.new_stock,
            transaction.performed_by.username if transaction.performed_by else 'System'
        ])
    
    wb.save(file_path)
    return file_path


def generate_event_resources_pdf(user, event_id=None):
    """Generate Event Resources PDF Report"""
    reports_dir = get_report_directory()
    filename = f'event_resources_{user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    file_path = os.path.join(reports_dir, filename)
    
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Event Resources Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    if event_id:
        try:
            event = Event.objects.get(pk=event_id)
            event_title = Paragraph(f"Event: {event.title}", styles['Heading2'])
            elements.append(event_title)
        except Event.DoesNotExist:
            pass
    
    # Get transactions related to events
    transactions = StockTransaction.objects.filter(reference_event_id=event_id) if event_id else StockTransaction.objects.filter(reference_event__isnull=False)
    
    data = [['Item', 'Quantity', 'Transaction Type', 'Date']]
    
    for transaction in transactions:
        data.append([
            transaction.inventory_item.name,
            str(transaction.quantity),
            transaction.get_transaction_type_display(),
            transaction.created_at.strftime('%Y-%m-%d')
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return file_path


def generate_event_resources_excel(user, event_id=None):
    """Generate Event Resources Excel Report"""
    reports_dir = get_report_directory()
    filename = f'event_resources_{user.id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    file_path = os.path.join(reports_dir, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Resources"
    
    headers = ['Item', 'SKU', 'Quantity', 'Transaction Type', 'Date', 'Event']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    transactions = StockTransaction.objects.filter(reference_event_id=event_id) if event_id else StockTransaction.objects.filter(reference_event__isnull=False)
    
    for transaction in transactions:
        ws.append([
            transaction.inventory_item.name,
            transaction.inventory_item.sku,
            transaction.quantity,
            transaction.get_transaction_type_display(),
            transaction.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            transaction.reference_event.title if transaction.reference_event else 'N/A'
        ])
    
    wb.save(file_path)
    return file_path
